from flask import Flask, render_template, request, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import cm
import io
import os

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL', 'sqlite:///ganadero.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# ================= MODELOS =================

class Config(db.Model):
    id    = db.Column(db.Integer, primary_key=True)
    clave = db.Column(db.String(50), unique=True)
    valor = db.Column(db.String(100))

class Animal(db.Model):
    id          = db.Column(db.Integer, primary_key=True)
    fecha       = db.Column(db.String(20))
    tipo        = db.Column(db.String(50))
    raza        = db.Column(db.String(50))
    peso        = db.Column(db.String(20))
    precio_kilo = db.Column(db.String(20))
    anos        = db.Column(db.String(10))
    importe     = db.Column(db.Float, default=0)
    comision    = db.Column(db.Float, default=0)
    transporte  = db.Column(db.Float, default=0)
    alimentacion= db.Column(db.Float, default=0)
    ayudante    = db.Column(db.Float, default=0)
    precio_total= db.Column(db.Float, default=0)
    precio_usd  = db.Column(db.Float, default=0)
    creado_en   = db.Column(db.DateTime, default=datetime.utcnow)

class Grupo(db.Model):
    id            = db.Column(db.Integer, primary_key=True)
    fecha         = db.Column(db.String(20))
    cantidad      = db.Column(db.Integer, default=1)
    tipo          = db.Column(db.String(50))
    raza          = db.Column(db.String(50))
    peso_total    = db.Column(db.String(20))
    pp            = db.Column(db.String(20))
    precio_kilo   = db.Column(db.String(20))
    anos          = db.Column(db.String(10))
    importe       = db.Column(db.Float, default=0)
    comision      = db.Column(db.Float, default=0)
    precio_liquido= db.Column(db.Float, default=0)
    transporte    = db.Column(db.Float, default=0)
    ayudante      = db.Column(db.Float, default=0)
    alimentacion  = db.Column(db.Float, default=0)
    precio_total  = db.Column(db.Float, default=0)
    precio_unidad = db.Column(db.Float, default=0)
    precio_usd    = db.Column(db.Float, default=0)
    creado_en     = db.Column(db.DateTime, default=datetime.utcnow)

# ================= HELPERS =================

def get_tc():
    cfg = Config.query.filter_by(clave='tipo_cambio').first()
    return float(cfg.valor) if cfg else 6.96

def fmt(v):
    try:
        return f"{float(v):,.2f}"
    except:
        return "0.00"

# ================= RUTAS PRINCIPALES =================

@app.route('/')
def index():
    tc = get_tc()
    return render_template('index.html', tipo_cambio=tc,
                           fecha_hoy=datetime.now().strftime('%d/%m/%Y'))

# ================= API CONFIG =================

@app.route('/api/config/tipo_cambio', methods=['GET'])
def get_tipo_cambio():
    return jsonify({'tipo_cambio': get_tc()})

@app.route('/api/config/tipo_cambio', methods=['POST'])
def set_tipo_cambio():
    data = request.json
    tc   = float(data.get('valor', 6.96))
    cfg  = Config.query.filter_by(clave='tipo_cambio').first()
    if cfg:
        cfg.valor = str(tc)
    else:
        db.session.add(Config(clave='tipo_cambio', valor=str(tc)))
    db.session.commit()
    return jsonify({'ok': True, 'tipo_cambio': tc})

# ================= API ANIMALES =================

@app.route('/api/animales', methods=['GET'])
def get_animales():
    animales = Animal.query.order_by(Animal.creado_en.desc()).all()
    return jsonify([{
        'id': a.id, 'fecha': a.fecha, 'tipo': a.tipo, 'raza': a.raza,
        'peso': a.peso, 'precio_kilo': a.precio_kilo, 'anos': a.anos,
        'importe': a.importe, 'comision': a.comision,
        'transporte': a.transporte, 'alimentacion': a.alimentacion,
        'ayudante': a.ayudante, 'precio_total': a.precio_total,
        'precio_usd': a.precio_usd
    } for a in animales])

@app.route('/api/animales', methods=['POST'])
def crear_animal():
    d  = request.json
    tc = get_tc()
    imp = float(d.get('importe', 0))
    com = float(d.get('comision', 0))
    tr  = float(d.get('transporte', 0))
    al  = float(d.get('alimentacion', 0))
    ay  = float(d.get('ayudante', 0))
    total = imp + com + tr + al + ay
    a = Animal(
        fecha=d.get('fecha'), tipo=d.get('tipo'), raza=d.get('raza'),
        peso=d.get('peso'), precio_kilo=d.get('precio_kilo'), anos=d.get('anos'),
        importe=imp, comision=com, transporte=tr, alimentacion=al, ayudante=ay,
        precio_total=round(total, 2), precio_usd=round(total / tc, 2)
    )
    db.session.add(a)
    db.session.commit()
    return jsonify({'ok': True, 'id': a.id})

@app.route('/api/animales/<int:id>', methods=['PUT'])
def editar_animal(id):
    a  = Animal.query.get_or_404(id)
    d  = request.json
    tc = get_tc()
    for campo in ['fecha','tipo','raza','peso','precio_kilo','anos']:
        setattr(a, campo, d.get(campo, getattr(a, campo)))
    imp = float(d.get('importe', 0))
    com = float(d.get('comision', 0))
    tr  = float(d.get('transporte', 0))
    al  = float(d.get('alimentacion', 0))
    ay  = float(d.get('ayudante', 0))
    total = imp + com + tr + al + ay
    a.importe=imp; a.comision=com; a.transporte=tr
    a.alimentacion=al; a.ayudante=ay
    a.precio_total=round(total,2); a.precio_usd=round(total/tc,2)
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/api/animales/<int:id>', methods=['DELETE'])
def eliminar_animal(id):
    a = Animal.query.get_or_404(id)
    db.session.delete(a)
    db.session.commit()
    return jsonify({'ok': True})

# ================= API GRUPOS =================

@app.route('/api/grupos', methods=['GET'])
def get_grupos():
    grupos = Grupo.query.order_by(Grupo.creado_en.desc()).all()
    return jsonify([{
        'id': g.id, 'fecha': g.fecha, 'cantidad': g.cantidad,
        'tipo': g.tipo, 'raza': g.raza, 'peso_total': g.peso_total,
        'pp': g.pp, 'precio_kilo': g.precio_kilo, 'anos': g.anos,
        'importe': g.importe, 'comision': g.comision,
        'precio_liquido': g.precio_liquido, 'transporte': g.transporte,
        'ayudante': g.ayudante, 'alimentacion': g.alimentacion,
        'precio_total': g.precio_total, 'precio_unidad': g.precio_unidad,
        'precio_usd': g.precio_usd
    } for g in grupos])

@app.route('/api/grupos', methods=['POST'])
def crear_grupo():
    d    = request.json
    tc   = get_tc()
    cant = int(d.get('cantidad', 1)) or 1
    imp  = float(d.get('importe', 0))
    com  = float(d.get('comision', 0))
    tr   = float(d.get('transporte', 0))
    ay   = float(d.get('ayudante', 0))
    al   = float(d.get('alimentacion', 0))
    liq  = imp + com
    total  = liq + tr + ay + al
    unidad = total / cant
    g = Grupo(
        fecha=d.get('fecha'), cantidad=cant, tipo=d.get('tipo'),
        raza=d.get('raza'), peso_total=d.get('peso_total'), pp=d.get('pp'),
        precio_kilo=d.get('precio_kilo'), anos=d.get('anos'),
        importe=imp, comision=com, precio_liquido=round(liq,2),
        transporte=tr, ayudante=ay, alimentacion=al,
        precio_total=round(total,2), precio_unidad=round(unidad,2),
        precio_usd=round(unidad/tc,2)
    )
    db.session.add(g)
    db.session.commit()
    return jsonify({'ok': True, 'id': g.id})

@app.route('/api/grupos/<int:id>', methods=['PUT'])
def editar_grupo(id):
    g  = Grupo.query.get_or_404(id)
    d  = request.json
    tc = get_tc()
    for campo in ['fecha','tipo','raza','peso_total','pp','precio_kilo','anos']:
        setattr(g, campo, d.get(campo, getattr(g, campo)))
    cant = int(d.get('cantidad', 1)) or 1
    imp  = float(d.get('importe', 0))
    com  = float(d.get('comision', 0))
    tr   = float(d.get('transporte', 0))
    ay   = float(d.get('ayudante', 0))
    al   = float(d.get('alimentacion', 0))
    liq  = imp + com
    total  = liq + tr + ay + al
    unidad = total / cant
    g.cantidad=cant; g.importe=imp; g.comision=com
    g.precio_liquido=round(liq,2); g.transporte=tr; g.ayudante=ay; g.alimentacion=al
    g.precio_total=round(total,2); g.precio_unidad=round(unidad,2)
    g.precio_usd=round(unidad/tc,2)
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/api/grupos/<int:id>', methods=['DELETE'])
def eliminar_grupo(id):
    g = Grupo.query.get_or_404(id)
    db.session.delete(g)
    db.session.commit()
    return jsonify({'ok': True})

# ================= EXPORTAR =================

def estilo_excel(ws, columnas, filas, titulo):
    fill_h = PatternFill("solid", fgColor="7C3AED")
    fill_s = PatternFill("solid", fgColor="5B21B6")
    fill_p = PatternFill("solid", fgColor="EDE9FE")
    borde  = Border(
        left=Side(style="thin",color="C4B5FD"), right=Side(style="thin",color="C4B5FD"),
        top=Side(style="thin",color="C4B5FD"),  bottom=Side(style="thin",color="C4B5FD")
    )
    centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ncols  = len(columnas)
    letra  = chr(64 + ncols)

    ws.merge_cells(f"A1:{letra}1")
    c = ws.cell(row=1, column=1,
                value=f"REPORTE: {titulo.upper()}  |  {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  USD: ${get_tc():.2f}")
    c.fill = fill_h; c.font = Font(bold=True,color="FFFFFF",size=13)
    c.alignment = centro; ws.row_dimensions[1].height = 30

    for ci, col in enumerate(columnas, 1):
        c = ws.cell(row=2, column=ci, value=col)
        c.fill = fill_s; c.font = Font(bold=True,color="FFFFFF",size=9)
        c.alignment = centro; c.border = borde
    ws.row_dimensions[2].height = 22

    from openpyxl.utils import get_column_letter
    for ri, fila in enumerate(filas, 3):
        for ci, val in enumerate(fila, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = Font(size=9); c.alignment = centro; c.border = borde
            if ri % 2 == 0: c.fill = fill_p
        ws.row_dimensions[ri].height = 18

    for ci in range(1, ncols + 1):
        letra_c = get_column_letter(ci)
        max_len = max((len(str(ws.cell(row=r, column=ci).value or ""))
                       for r in range(1, len(filas)+4)), default=8)
        ws.column_dimensions[letra_c].width = max(max_len + 4, 12)

    if filas:
        ft = ws.max_row + 1
        ws.merge_cells(f"A{ft}:{chr(64+ncols)}{ft}")
        c = ws.cell(row=ft, column=1, value=f"Total registros: {len(filas)}")
        c.fill = fill_h; c.font = Font(bold=True,color="FFFFFF",size=9)
        c.alignment = centro

@app.route('/exportar/animales/excel')
def exportar_animales_excel():
    animales = Animal.query.order_by(Animal.creado_en.desc()).all()
    cols  = ["Fecha","Tipo","Raza","Peso","Precio Kilo","Anios",
             "Importe","Comision","Transporte","Alimentacion","Ayudante",
             "Precio Total","Precio USD"]
    filas = [[a.fecha,a.tipo,a.raza,a.peso,a.precio_kilo,a.anos,
              a.importe,a.comision,a.transporte,a.alimentacion,a.ayudante,
              a.precio_total,a.precio_usd] for a in animales]
    wb = Workbook(); ws = wb.active; ws.title = "Animales"
    estilo_excel(ws, cols, filas, "Animales")
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=f"animales_{datetime.now().strftime('%Y%m%d')}.xlsx")

@app.route('/exportar/grupos/excel')
def exportar_grupos_excel():
    grupos = Grupo.query.order_by(Grupo.creado_en.desc()).all()
    cols  = ["Fecha","Cantidad","Tipo","Raza","Peso Total","P.P.",
             "Precio Kilo","Anios","Importe","Comision","Precio Liquido",
             "Transporte","Ayudante","Alimentacion",
             "Precio Total","Precio Unidad","Precio USD"]
    filas = [[g.fecha,g.cantidad,g.tipo,g.raza,g.peso_total,g.pp,
              g.precio_kilo,g.anos,g.importe,g.comision,g.precio_liquido,
              g.transporte,g.ayudante,g.alimentacion,
              g.precio_total,g.precio_unidad,g.precio_usd] for g in grupos]
    wb = Workbook(); ws = wb.active; ws.title = "Grupos"
    estilo_excel(ws, cols, filas, "Grupos")
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=f"grupos_{datetime.now().strftime('%Y%m%d')}.xlsx")

def exportar_pdf_tabla(titulo, columnas, filas):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.8*cm, bottomMargin=1.2*cm)
    estilos = getSampleStyleSheet()
    story   = [
        Paragraph(f"Reporte: {titulo.upper()}", ParagraphStyle(
            "t", parent=estilos["Title"],
            fontSize=15, textColor=colors.HexColor("#7c3aed"), spaceAfter=4)),
        Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  USD: ${get_tc():.2f}",
                  ParagraphStyle("s", parent=estilos["Normal"],
                                 fontSize=8, textColor=colors.grey, spaceAfter=14))
    ]
    if filas:
        ancho = (landscape(A4)[0] - 3*cm) / len(columnas)
        data  = [columnas] + filas
        t = Table(data, colWidths=[ancho]*len(columnas), repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND",    (0,0),(-1,0),  colors.HexColor("#7c3aed")),
            ("TEXTCOLOR",     (0,0),(-1,0),  colors.white),
            ("FONTNAME",      (0,0),(-1,0),  "Helvetica-Bold"),
            ("FONTSIZE",      (0,0),(-1,-1), 7),
            ("ALIGN",         (0,0),(-1,-1), "CENTER"),
            ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
            ("TOPPADDING",    (0,0),(-1,-1), 4),
            ("BOTTOMPADDING", (0,0),(-1,-1), 4),
            ("ROWBACKGROUNDS",(0,1),(-1,-1), [colors.white, colors.HexColor("#ede9fe")]),
            ("GRID",          (0,0),(-1,-1), 0.4, colors.HexColor("#c4b5fd")),
            ("LINEBELOW",     (0,0),(-1,0),  1.5, colors.HexColor("#5b21b6")),
        ]))
        story.append(t)
    story.append(Spacer(1, 0.4*cm))
    story.append(Paragraph("Sistema Ganadero Profesional",
                            ParagraphStyle("p", parent=estilos["Normal"],
                                           fontSize=7, textColor=colors.grey, alignment=1)))
    doc.build(story)
    buf.seek(0)
    return buf

@app.route('/exportar/animales/pdf')
def exportar_animales_pdf():
    animales = Animal.query.order_by(Animal.creado_en.desc()).all()
    cols  = ["Fecha","Tipo","Raza","Peso","P.Kilo","Anios",
             "Importe","Comision","Transporte","Alimento","Ayudante","Total","USD"]
    filas = [[a.fecha,a.tipo,a.raza,a.peso,a.precio_kilo,a.anos,
              a.importe,a.comision,a.transporte,a.alimentacion,a.ayudante,
              a.precio_total,a.precio_usd] for a in animales]
    buf = exportar_pdf_tabla("Animales", cols, filas)
    return send_file(buf, mimetype='application/pdf', as_attachment=True,
                     download_name=f"animales_{datetime.now().strftime('%Y%m%d')}.pdf")

@app.route('/exportar/grupos/pdf')
def exportar_grupos_pdf():
    grupos = Grupo.query.order_by(Grupo.creado_en.desc()).all()
    cols  = ["Fecha","Cant","Tipo","Raza","Peso","P.P.",
             "P.Kilo","Anios","Importe","Comision","P.Liquido",
             "Transporte","Ayudante","Alimento","Total","Unidad","USD"]
    filas = [[g.fecha,g.cantidad,g.tipo,g.raza,g.peso_total,g.pp,
              g.precio_kilo,g.anos,g.importe,g.comision,g.precio_liquido,
              g.transporte,g.ayudante,g.alimentacion,
              g.precio_total,g.precio_unidad,g.precio_usd] for g in grupos]
    buf = exportar_pdf_tabla("Grupos", cols, filas)
    return send_file(buf, mimetype='application/pdf', as_attachment=True,
                     download_name=f"grupos_{datetime.now().strftime('%Y%m%d')}.pdf")

# ================= INIT =================

with app.app_context():
    db.create_all()

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
